from openpyxl import Workbook
import datetime

wb = Workbook()

ws = wb.active

ws['A1'] = 42

ws.append([1, 2, 3])

ws['A2'] = datetime.datetime.now()
ws['A4'] = 57

wb.save("sample.xlsx")